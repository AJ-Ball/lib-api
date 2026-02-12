from __future__ import annotations

import re
from typing import Optional, Dict, Any, List, Tuple

from fastapi import FastAPI, Query
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook

EXCEL_PATH = "Data_Lib.xlsx"
SHEET_NAME = "api"

app = FastAPI(title="Library Locator API", version="1.1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # ถ้าจะเข้มงวดค่อยจำกัดโดเมนทีหลัง
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# cache data in memory
_rows: Optional[List[Dict[str, Any]]] = None

THAI_SUFFIX_RE = re.compile(r"(?P<num>[0-9.]+)\s*(?P<suffix>[ก-๙]{0,3})$", re.UNICODE)


def to_float(x) -> Optional[float]:
    if x is None:
        return None
    try:
        return float(x)
    except Exception:
        return None


def to_int(x) -> Optional[int]:
    if x is None:
        return None
    try:
        return int(float(x))
    except Exception:
        return None


def clean_suffix(x: str) -> str:
    """เอาเฉพาะตัวอักษรไทย 1–3 ตัว กันเคสมี '-' '.' ช่องว่าง ฯลฯ"""
    if not x:
        return ""
    m = re.search(r"[ก-๙]{1,3}", str(x))
    return m.group(0) if m else ""


def num_to_key(x) -> Optional[int]:
    """
    แปลงเลขหมวดให้เป็น key แบบ 3+3 หลัก:
      - 370.113 -> 370113
      - 370113  -> 370113
      - "370.113" -> 370113
      - "370113พ" -> 370113
    """
    if x is None:
        return None

    # numeric
    if isinstance(x, (int, float)):
        v = float(x)
        if v < 1000:
            return int(round(v * 1000))
        return int(round(v))

    # string
    s = str(x).strip()
    m = re.search(r"([0-9]+(?:\.[0-9]+)?)", s)
    if not m:
        return None

    v = float(m.group(1))
    if v < 1000:
        return int(round(v * 1000))
    return int(round(v))


def normalize_call_number(raw: str) -> Tuple[Optional[int], str]:
    """
    รับอินพุต:
      - "370.113"
      - "370113"
      - "370.113พ"
      - "370113พ"
    คืนค่า:
      (q_key, q_suffix)
      q_key = 370113
      q_suffix = 'พ'
    """
    if not raw:
        return None, ""

    s = str(raw).strip().replace("–", "-").replace("—", "-")
    m = THAI_SUFFIX_RE.search(s)
    if not m:
        # ถ้าไม่ match รูปแบบ suffix ก็ยังพยายามดึงเลขอย่างเดียว
        key = num_to_key(s)
        return key, ""

    num_part = m.group("num").strip()
    suffix = clean_suffix(m.group("suffix") or "")

    # ถ้าไม่มีจุด ให้ใส่จุดหลัง 3 หลักแรก
    if "." not in num_part:
        digits = re.sub(r"\D", "", num_part)
        if len(digits) <= 3:
            # เช่น "370" => 370000
            try:
                key = int(digits) * 1000
                return key, suffix
            except Exception:
                return None, suffix
        num_part = digits[:3] + "." + digits[3:]

    try:
        num_val = float(num_part)
    except Exception:
        return None, suffix

    key = num_to_key(num_val)
    return key, suffix


def load_data() -> List[Dict[str, Any]]:
    """Load Excel sheet to memory once."""
    global _rows
    if _rows is not None:
        return _rows

    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]

    # header row
    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [(str(c).strip() if c is not None else "") for c in header_cells]

    rows: List[Dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        d: Dict[str, Any] = {}
        for k, v in zip(headers, r):
            if k:
                d[k] = v

        # normalize numeric-like columns if exist
        d["range_start_num"] = to_float(d.get("range_start_num"))
        d["range_end_num"] = to_float(d.get("range_end_num"))

        d["row"] = to_int(d.get("row"))
        d["shelf_level"] = to_int(d.get("shelf_level"))
        d["locker"] = to_int(d.get("locker"))
        d["building_floor"] = to_int(d.get("building_floor"))

        # strings
        for c in [
            "id", "side", "category", "call_range", "map_url",
            "range_start_raw", "range_end_raw",
            "range_start_suffix", "range_end_suffix",
        ]:
            d[c] = (str(d[c]).strip() if d.get(c) is not None else "")

        # build keys (ใช้ raw เป็น fallback)
        d["start_key"] = num_to_key(d.get("range_start_num") if d.get("range_start_num") is not None else d.get("range_start_raw"))
        d["end_key"] = num_to_key(d.get("range_end_num") if d.get("range_end_num") is not None else d.get("range_end_raw"))

        # clean suffix
        d["range_start_suffix_clean"] = clean_suffix(d.get("range_start_suffix", ""))
        d["range_end_suffix_clean"] = clean_suffix(d.get("range_end_suffix", ""))

        # drop invalid
        if d["start_key"] is None or d["end_key"] is None:
            continue

        rows.append(d)

    wb.close()
    _rows = rows
    return _rows


def match_row(d: Dict[str, Any], q_key: int, q_suffix: str, strict_suffix: bool) -> bool:
    start_k = d["start_key"]
    end_k = d["end_key"]

    if q_key < start_k or q_key > end_k:
        return False

    if not strict_suffix:
        return True

    # กลางช่วงผ่านเลย
    if q_key > start_k and q_key < end_k:
        return True

    start_s = d.get("range_start_suffix_clean", "") or ""
    end_s = d.get("range_end_suffix_clean", "") or ""
    q_suffix = clean_suffix(q_suffix)

    # ชนขอบต้น
    if q_key == start_k and start_s:
        if q_suffix < start_s:
            return False

    # ชนขอบท้าย
    if q_key == end_k and end_s:
        if q_suffix > end_s:
            return False

    return True


def span_key(d: Dict[str, Any]) -> int:
    return abs(int(d["end_key"]) - int(d["start_key"]))


def rank_candidates(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    จัดอันดับ:
      1) ช่วงแคบสุดก่อน (ลดความกำกวม)
      2) row / shelf_level / locker (กันสุ่ม)
    """
    return sorted(
        items,
        key=lambda d: (
            span_key(d),
            d.get("row") if d.get("row") is not None else 10**9,
            d.get("shelf_level") if d.get("shelf_level") is not None else 10**9,
            d.get("locker") if d.get("locker") is not None else 10**9,
        ),
    )


@app.get("/")
def root() -> Dict[str, Any]:
    return {
        "service": "Library Locator API",
        "endpoints": ["/health", "/search", "/docs"],
    }


@app.get("/health")
def health() -> Dict[str, Any]:
    data = load_data()
    return {"ok": True, "rows": len(data)}


@app.get("/search")
def search(
    q: str = Query(..., description="เลขหมวดหรือคำค้น เช่น 370.113พ หรือ 370113 หรือ สังคมศาสตร์"),
    limit: int = Query(5, ge=1, le=20),
) -> Dict[str, Any]:
    data = load_data()
    q = (q or "").strip()

    # 1) Try as call number
    q_key, q_suffix = normalize_call_number(q)
    if q_key is not None:
        strict_suffix = bool(clean_suffix(q_suffix))
        hits = [d for d in data if match_row(d, q_key, q_suffix, strict_suffix)]

        if not hits:
            return {
                "found": False,
                "mode": "call_number",
                "query": q,
                "normalized": {"key": q_key, "suffix": clean_suffix(q_suffix)},
                "results": [],
                "suggest": [
                    "ลองพิมพ์เฉพาะตัวเลข เช่น 370.113 หรือ 370113",
                    "ถ้ามีตัวอักษรไทยท้ายเลข ให้ใส่ด้วย เช่น 370.113พ",
                    "หรือค้นด้วยชื่อหมวด เช่น สังคมศาสตร์",
                ],
            }

        hits = rank_candidates(hits)[:limit]
        results = []
        for d in hits:
            results.append(
                {
                    "id": d.get("id", ""),
                    "call_range": d.get("call_range", ""),
                    "category": d.get("category", ""),
                    "location": {
                        "row": d.get("row"),
                        "shelf_level": d.get("shelf_level"),
                        "locker": d.get("locker"),
                        "building_floor": d.get("building_floor"),
                        "side": d.get("side", ""),
                    },
                    "range": {
                        "start_raw": d.get("range_start_raw", ""),
                        "end_raw": d.get("range_end_raw", ""),
                        "start_key": d.get("start_key"),
                        "end_key": d.get("end_key"),
                        "start_suffix": d.get("range_start_suffix_clean", ""),
                        "end_suffix": d.get("range_end_suffix_clean", ""),
                    },
                    "map_url": d.get("map_url", ""),
                }
            )

        # ช่วย Jotform: มีข้อความสรุปพร้อมใช้
        top = results[0]
        loc = top["location"]
        message = (
            f"พบหนังสือที่ แถว {loc.get('row')} ชั้น {loc.get('shelf_level')} "
            f"ล็อค {loc.get('locker')} อาคารชั้น {loc.get('building_floor')} "
            f"(ด้าน{loc.get('side')})"
        )

        return {
            "found": True,
            "mode": "call_number",
            "query": q,
            "normalized": {"key": q_key, "suffix": clean_suffix(q_suffix)},
            "count": len(results),
            "message": message,
            "results": results,
        }

    # 2) Text search
    q_low = q.lower()

    def s(x) -> str:
        return (x or "").lower()

    hits = [
        d for d in data
        if (q_low in s(d.get("category")) or q_low in s(d.get("call_range")) or q_low in s(d.get("id")))
    ]

    if not hits:
        return {
            "found": False,
            "mode": "text",
            "query": q,
            "results": [],
            "suggest": [
                "ลองค้นด้วยเลขหมวด เช่น 370.113 หรือ 370113",
                "หรือค้นด้วยคำหมวด เช่น สังคมศาสตร์ / วิทยาศาสตร์",
            ],
        }

    hits = hits[:limit]
    results = []
    for d in hits:
        results.append(
            {
                "id": d.get("id", ""),
                "call_range": d.get("call_range", ""),
                "category": d.get("category", ""),
                "location": {
                    "row": d.get("row"),
                    "shelf_level": d.get("shelf_level"),
                    "locker": d.get("locker"),
                    "building_floor": d.get("building_floor"),
                    "side": d.get("side", ""),
                },
                "map_url": d.get("map_url", ""),
            }
        )

    top = results[0]
    loc = top["location"]
    message = (
        f"พบหมวด '{top.get('category')}' ที่ แถว {loc.get('row')} ชั้น {loc.get('shelf_level')} "
        f"ล็อค {loc.get('locker')} อาคารชั้น {loc.get('building_floor')} (ด้าน{loc.get('side')})"
    )

    return {
        "found": True,
        "mode": "text",
        "query": q,
        "count": len(results),
        "message": message,
        "results": results,
    }
